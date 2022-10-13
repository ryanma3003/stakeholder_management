from django.shortcuts import redirect
from django.urls import reverse_lazy, reverse
from django.views.generic import ListView, DetailView, CreateView, DeleteView, UpdateView
from django.contrib import messages
from django.http import HttpResponseRedirect, FileResponse
import openpyxl
import os
from django.conf import settings

# Create your views here.

from .models import Tmpi
from .forms import TmpiForm
from phrasalword.views import decrypt as decrypt_aes
from phrasalword.models import Phrasalword

import cryptography
from cryptography.fernet import Fernet
from cryptography.hazmat.primitives.kdf.scrypt import Scrypt

import secrets
import base64
import tempfile
import hashlib

__key__ = hashlib.sha256(settings.AES_KEY).digest()
__dir__ = 'C:\\xampp\\htdocs\\csirt-py\\csirt\\media\\ttis_evaluasi'

def generate_salt(size=16):
    """Generate the salt used for key derivation, 
    `size` is the length of the salt to generate"""
    return secrets.token_bytes(size)

def derive_key(salt, password):
    """Derive the key from the `password` using the passed `salt`"""
    kdf = Scrypt(salt=salt, length=32, n=2**14, r=8, p=1)
    return kdf.derive(password.encode())

def load_salt():
    # load salt from salt.salt file
    return open(os.path.join(__dir__, "salt.salt"), "rb").read()

def generate_key(password, salt_size=16, load_existing_salt=False, save_salt=True):
    """
    Generates a key from a `password` and the salt.
    If `load_existing_salt` is True, it'll load the salt from a file
    in the current directory called "salt.salt".
    If `save_salt` is True, then it will generate a new salt
    and save it to "salt.salt"
    """
    if load_existing_salt:
        # load existing salt
        salt = load_salt()
    elif save_salt:
        # generate new salt and save it
        salt = generate_salt(salt_size)
        with open(os.path.join(__dir__, "salt.salt"), "wb") as salt_file:
            salt_file.write(salt)
    # generate the key from the salt and the password
    derived_key = derive_key(salt, password)
    # encode it using Base 64 and return it
    return base64.urlsafe_b64encode(derived_key)

def encrypt_file(filename, key):
    """
    Given a filename (str) and key (bytes), it encrypts the file and write it
    """
    f = Fernet(key)
    with open(filename, "rb") as file:
        # read all file data
        file_data = file.read()
    # encrypt data
    encrypted_data = f.encrypt(file_data)
    # write the encrypted file
    with open(filename, "wb") as file:
        file.write(encrypted_data)

def decrypt_file(filename, key):
    """
    Given a filename (str) and key (bytes), it decrypts the file and write it
    """
    f = Fernet(key)
    with open(filename, "rb") as file:
        # read the encrypted data
        encrypted_data = file.read()
    # decrypt data
    try:
        decrypted_data = f.decrypt(encrypted_data)
    except cryptography.fernet.InvalidToken:
        print("Invalid token, most likely the password is incorrect")
        return
    # write the original file
    tmp = tempfile.NamedTemporaryFile(delete=False)
    with open(tmp.name, "wb") as file:
        file.write(decrypted_data)
    print("File decrypted successfully")

def downloadExcel(request, show_id):
    query = Tmpi.objects.get(id = show_id)
    file_tmpi = query.file_tmpi

    password = request.POST.get('passphrase')

    phrasalword = Phrasalword.objects.get(user_id = request.user.id).passphrase
    dec_pass = decrypt_aes(phrasalword)
    key = generate_key(password, load_existing_salt=True)

    if password == dec_pass:
        try: 
            # decrypt_file(os.path.join(__dir__, file_tmpi.name.split("/")[-1]), key)
            f = Fernet(key)
            with open(os.path.join(__dir__, file_tmpi.name.split("/")[-1]), "rb") as file:
                encrypted_data = file.read()
            try:
                decrypted_data = f.decrypt(encrypted_data)
            except cryptography.fernet.InvalidToken:
                print("Invalid token, most likely the password is incorrect")
                return
            tmp = tempfile.NamedTemporaryFile(delete=False)
            with open(tmp.name, "wb") as file:
                file.write(decrypted_data)
            return FileResponse(open(tmp.name, 'rb'), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        except FileNotFoundError:
            messages.error(request, "File not found.")
            return HttpResponseRedirect(reverse('stakeholder:detail', kwargs={'pk':request.GET.get('s_id')} ))
    else :
        messages.error(request, "Wrong passphrase.")
        return HttpResponseRedirect(reverse('stakeholder:detail', kwargs={'pk':request.GET.get('s_id')} ))

class SearchList():
    # LoginRequiredMixin
    # login_url = '/login'
    # redirect_field_name = 'redirect_to'

    def get_list_data(self, get_request):
        if len(get_request) == 0:
            search_result = Tmpi.objects.all()
        elif get_request.__contains__('filter'):
            search_result = Tmpi.objects.filter(stakeholder=get_request['filter'])
        else :
            search_result = Tmpi.objects.none()
        return search_result

class TmpiListView(SearchList, ListView):
    model = Tmpi
    context_object_name = 'tmpi_list'
    ordering = ['stakeholder']
    # paginate_by: 3

    def get_queryset(self):
        self.queryset = self.get_list_data(self.request.GET)
        return super(TmpiListView, self).get_queryset()

    def get_context_data(self, *args, **kwargs):
        list_stakeholder = self.model.objects.values('stakeholder', 'stakeholder__name').distinct()
        
        context = super(TmpiListView, self).get_context_data(*args, **kwargs)
        context['title'] = 'Pengukuran TMPI'
        context['list_stakeholder'] = list_stakeholder
        return context

class TmpiDetailView(DetailView):
    model = Tmpi
    extra_context = {}

    def get_context_data(self, *args, **kwargs):

        context = super(TmpiDetailView, self).get_context_data(*args, **kwargs)
        context['title'] = "%s %s" % (self.object.stakeholder, 'Tmpi')
        
        return context

class TmpiCreateView(CreateView):
    form_class = TmpiForm
    template_name = 'tmpi/tmpi_create.html'
    extra_context = {
        'title' : 'Create TMPI',
        'breadcrumb': 'Create'
    }

    def get_context_data(self, *args, **kwargs):
        kwargs.update(self.extra_context)
        return super(TmpiCreateView, self).get_context_data(*args, **kwargs)

    def form_valid(self, form):
        # file_tmpi = form.cleaned_data.get('file_tmpi', False)
        file_tmpi = self.request.FILES.get('file_tmpi')

        if file_tmpi :
            wb = openpyxl.load_workbook(file_tmpi, data_only=True)

            # getting a particular sheet by name out of many sheets
            worksheet = wb["Hasil Perhitungan Tk. Maturitas"]

            # excel_data = list()
            # # iterating over the rows and
            # # getting value from each cell in row
            # for row in worksheet.iter_rows():
            #     row_data = list()
            #     for cell in row:
            #         row_data.append(str(cell.value))
            #     excel_data.append(row_data)

            form.instance.penilaian_kritikalitas = worksheet["AA4"].value if worksheet["AA4"].value else worksheet["Z4"].value
            form.instance.analisis_ancaman = worksheet["AA5"].value if worksheet["AA5"].value else worksheet["Z5"].value
            form.instance.orang_proses_teknologi = worksheet["AA6"].value if worksheet["AA6"].value else worksheet["Z6"].value
            form.instance.lingkungan_kontrol = worksheet["AA7"].value if worksheet["AA7"].value else worksheet["Z7"].value
            form.instance.penilaian_kematangan = worksheet["AA8"].value if worksheet["AA8"].value else worksheet["Z8"].value
            form.instance.total_fase_1 = worksheet["AA9"].value if worksheet["AA9"].value else worksheet["Z9"].value

            form.instance.identifikasi_respon = worksheet["AA10"].value if worksheet["AA10"].value else worksheet["Z10"].value
            form.instance.penyelidikan = worksheet["AA11"].value if worksheet["AA11"].value else worksheet["Z11"].value
            form.instance.aksi = worksheet["AA12"].value if worksheet["AA12"].value else worksheet["Z12"].value
            form.instance.pemulihan = worksheet["AA13"].value if worksheet["AA13"].value else worksheet["Z13"].value
            form.instance.total_fase_2 = worksheet["AA14"].value if worksheet["AA14"].value else worksheet["Z14"].value

            form.instance.identifikasi_tindak_lanjut = worksheet["AA15"].value if worksheet["AA15"].value else worksheet["Z15"].value
            form.instance.pelaporan_review = worksheet["AA16"].value if worksheet["AA16"].value else worksheet["Z16"].value
            form.instance.pembelajaran = worksheet["AA17"].value if worksheet["AA17"].value else worksheet["Z17"].value
            form.instance.pembaruan_informasi = worksheet["AA18"].value if worksheet["AA18"].value else worksheet["Z18"].value
            form.instance.analisis_tren = worksheet["AA19"].value if worksheet["AA19"].value else worksheet["Z19"].value
            form.instance.total_fase_3 = worksheet["AA20"].value if worksheet["AA20"].value else worksheet["Z20"].value

            form.instance.nilai_akhir = worksheet["AA22"].value if worksheet["AA22"].value else worksheet["Z22"].value

            tmpi = form.save(commit=False)
            tmpi.file_tmpi = file_tmpi
            tmpi.save()

            directory = 'C:\\xampp\\htdocs\\csirt-py\\csirt\\media\\ttis_evaluasi'
            phrasalword = Phrasalword.objects.get(user_id = self.request.user.id).passphrase
            password = decrypt_aes(phrasalword)

            key = generate_key(password, load_existing_salt=True)
            encrypt_file(os.path.join(directory, file_tmpi.name), key)

        # return redirect('tmpi:detail', self.object.id)
        return super().form_valid(form)

class TmpiUpdateView(UpdateView):
    model = Tmpi
    form_class = TmpiForm
    template_name = 'tmpi/tmpi_create.html'
    extra_context = {
        'title' : 'Update TMPI',
        'breadcrumb': 'Update'
    }

    def get_context_data(self, *args, **kwargs):
        kwargs.update(self.extra_context)
        return super(TmpiUpdateView, self).get_context_data(*args, **kwargs)

    def form_valid(self, form):
        file_tmpi = self.request.FILES.get('file_tmpi')

        if file_tmpi :
            wb = openpyxl.load_workbook(file_tmpi, data_only=True)

            # getting a particular sheet by name out of many sheets
            worksheet = wb["Hasil Perhitungan Tk. Maturitas"]

            form.instance.penilaian_kritikalitas = worksheet["AA4"].value if worksheet["AA4"].value else worksheet["Z4"].value
            form.instance.analisis_ancaman = worksheet["AA5"].value if worksheet["AA5"].value else worksheet["Z5"].value
            form.instance.orang_proses_teknologi = worksheet["AA6"].value if worksheet["AA6"].value else worksheet["Z6"].value
            form.instance.lingkungan_kontrol = worksheet["AA7"].value if worksheet["AA7"].value else worksheet["Z7"].value
            form.instance.penilaian_kematangan = worksheet["AA8"].value if worksheet["AA8"].value else worksheet["Z8"].value
            form.instance.total_fase_1 = worksheet["AA9"].value if worksheet["AA9"].value else worksheet["Z9"].value

            form.instance.identifikasi_respon = worksheet["AA10"].value if worksheet["AA10"].value else worksheet["Z10"].value
            form.instance.penyelidikan = worksheet["AA11"].value if worksheet["AA11"].value else worksheet["Z11"].value
            form.instance.aksi = worksheet["AA12"].value if worksheet["AA12"].value else worksheet["Z12"].value
            form.instance.pemulihan = worksheet["AA13"].value if worksheet["AA13"].value else worksheet["Z13"].value
            form.instance.total_fase_2 = worksheet["AA14"].value if worksheet["AA14"].value else worksheet["Z14"].value

            form.instance.identifikasi_tindak_lanjut = worksheet["AA15"].value if worksheet["AA15"].value else worksheet["Z15"].value
            form.instance.pelaporan_review = worksheet["AA16"].value if worksheet["AA16"].value else worksheet["Z16"].value
            form.instance.pembelajaran = worksheet["AA17"].value if worksheet["AA17"].value else worksheet["Z17"].value
            form.instance.pembaruan_informasi = worksheet["AA18"].value if worksheet["AA18"].value else worksheet["Z18"].value
            form.instance.analisis_tren = worksheet["AA19"].value if worksheet["AA19"].value else worksheet["Z19"].value
            form.instance.total_fase_3 = worksheet["AA20"].value if worksheet["AA20"].value else worksheet["Z20"].value

            form.instance.nilai_akhir = worksheet["AA22"].value if worksheet["AA22"].value else worksheet["Z22"].value

            tmpi = form.save(commit=False)
            tmpi.file_tmpi = file_tmpi
            tmpi.save()

            directory = 'C:\\xampp\\htdocs\\csirt-py\\csirt\\media\\ttis_evaluasi'
            phrasalword = Phrasalword.objects.get(user_id = self.request.user.id).passphrase
            password = decrypt_aes(phrasalword)

            # key = generate_key(password, salt_size=16, save_salt=True)
            key = generate_key(password, load_existing_salt=True)
            encrypt_file(os.path.join(directory, file_tmpi.name), key)

        return super().form_valid(form)

class TmpiDeleteView(DeleteView):
    model = Tmpi
    success_url = reverse_lazy('tmpi:index')
